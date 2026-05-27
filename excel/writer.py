import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

EXCEL_PATH = "data/solicitud_inscripciones_DA.xlsx"


def marcar_como_procesado(indices):

    wb = load_workbook(EXCEL_PATH)

    ws = wb["Solicitudes"]

    for idx in indices:

        fila_excel = idx + 2

        ws[f"F{fila_excel}"] = "SI"

    wb.save(EXCEL_PATH)


def agregar_reporte(reporte_csv):

    df_reporte = pd.read_csv(reporte_csv)

    wb = load_workbook(EXCEL_PATH)

    ws = wb["Reportes"]

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for _, row in df_reporte.iterrows():

        ws.append([
            fecha,
            row["course_id"],
            row["user_id"],
            row["role"],
            row["status"],
            row["register_status"],
            row["observations"]
        ])

    wb.save(EXCEL_PATH)